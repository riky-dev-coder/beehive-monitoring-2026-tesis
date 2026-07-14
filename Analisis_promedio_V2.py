import pandas as pd
import numpy as np
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ─── Configuración ────────────────────────────────────────────────────────────
INPUT_CSV  = "feeds(1).csv"
OUTPUT_XLS = "analisis_promedios.xlsx"

# Mapeo de campos por cámara
CAMARAS = {
    "Cría": {
        "temperatura": "field1",
        "humedad":     "field4",
        "peso":        "field7",
    },
    "Mielera": {
        "temperatura": "field2",
        "humedad":     "field5",
        "peso":        "field8",
    },
    "Exterior": {
        "temperatura": "field3",
        "humedad":     "field6",
        # sin peso para exterior
    },
}

# ─── Estilos ──────────────────────────────────────────────────────────────────
PALETA = {
    "Cría":     {"header": "1F4E79", "subheader": "2E75B6", "alt": "D6E4F0"},
    "Mielera":  {"header": "375623", "subheader": "548235", "alt": "E2EFDA"},
    "Exterior": {"header": "7B3F00", "subheader": "C0602A", "alt": "FCE4D6"},
    "General":  {"header": "3A3A3A", "subheader": "595959", "alt": "EDEDED"},
}

thin   = Side(style="thin", color="BFBFBF")
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)
CENTER = Alignment(horizontal="center", vertical="center")

def make_fill(color):  return PatternFill("solid", start_color=color, end_color=color)
def make_font(color="000000", bold=False, size=10):
    return Font(name="Arial", bold=bold, color=color, size=size)

# ─── Carga y limpieza ─────────────────────────────────────────────────────────
df_raw = pd.read_csv(INPUT_CSV, parse_dates=["created_at"])
df_raw["created_at"] = pd.to_datetime(df_raw["created_at"], utc=True).dt.tz_localize(None)

all_fields = {v for cam in CAMARAS.values() for v in cam.values()}
for col in all_fields:
    df_raw[col] = df_raw[col].replace(0, np.nan)

df_raw["semana"]   = df_raw["created_at"].dt.to_period("W")
df_raw["quincena"] = df_raw["created_at"].apply(
    lambda x: f"{x.year}-{x.month:02d}-{'Q1 (1-15)' if x.day <= 15 else 'Q2 (16-fin)'}"
)
df_raw["mes"] = df_raw["created_at"].dt.to_period("M")

# ─── Cálculo de promedios ─────────────────────────────────────────────────────
UNITS = {"temperatura": "°C", "humedad": "%", "peso": "kg"}

def compute_summaries(fields: dict, period_col: str, period_label: str) -> pd.DataFrame:
    metric_cols   = list(fields.values())
    display_names = {v: k for k, v in fields.items()}

    grp = df_raw.groupby(period_col)[metric_cols].agg(["mean", "count"])
    grp.columns = [f"{display_names[col]}_{stat}" for col, stat in grp.columns]
    grp = grp.reset_index()
    grp[period_col] = grp[period_col].astype(str)
    grp = grp.rename(columns={period_col: period_label})

    mean_cols = [c for c in grp.columns if c.endswith("_mean")]
    grp[mean_cols] = grp[mean_cols].round(2)

    rename = {}
    for metric in fields:
        unit = UNITS.get(metric, "")
        rename[f"{metric}_mean"]  = f"{metric.capitalize()} Prom. ({unit})"
        rename[f"{metric}_count"] = f"N ({metric})"
    grp = grp.rename(columns=rename)

    ordered = [period_label]
    for metric in fields:
        unit = UNITS.get(metric, "")
        ordered += [f"{metric.capitalize()} Prom. ({unit})", f"N ({metric})"]
    return grp[ordered]

summaries = {}
for camara, fields in CAMARAS.items():
    summaries[camara] = {
        "Semana":   compute_summaries(fields, "semana",   "Semana"),
        "Quincena": compute_summaries(fields, "quincena", "Quincena"),
        "Mes":      compute_summaries(fields, "mes",      "Mes"),
    }

# ─── Escritura Excel ──────────────────────────────────────────────────────────
def write_table(ws, title, data, start_row, palette):
    n_cols = len(data.columns)

    ws.merge_cells(
        start_row=start_row, start_column=1,
        end_row=start_row,   end_column=n_cols,
    )
    cell = ws.cell(row=start_row, column=1, value=title)
    cell.font, cell.fill, cell.alignment = (
        make_font("FFFFFF", bold=True, size=11),
        make_fill(palette["header"]),
        CENTER,
    )
    start_row += 1

    for ci, col_name in enumerate(data.columns, 1):
        cell = ws.cell(row=start_row, column=ci, value=col_name)
        cell.font, cell.fill, cell.alignment, cell.border = (
            make_font("FFFFFF", bold=True, size=10),
            make_fill(palette["subheader"]),
            CENTER, BORDER,
        )
    start_row += 1

    white = make_fill("FFFFFF")
    alt   = make_fill(palette["alt"])
    for ri, (_, row) in enumerate(data.iterrows()):
        fill = alt if ri % 2 == 0 else white
        for ci, val in enumerate(row, 1):
            cell = ws.cell(row=start_row + ri, column=ci, value=val)
            cell.font, cell.fill, cell.alignment, cell.border = (
                make_font(), fill, CENTER, BORDER,
            )
    return start_row + len(data) + 2


def build_camera_sheet(wb, camara, s_week, s_quin, s_month):
    ws = wb.create_sheet(f"Cámara {camara}")
    ws.sheet_view.showGridLines = False
    palette = PALETA[camara]

    n_cols = len(s_week.columns)
    widths = [24] + [20, 13] * ((n_cols - 1) // 2 + 1)
    for i, w in enumerate(widths[:n_cols], 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    r = 2
    r = write_table(ws, f"📅 {camara.upper()} — PROMEDIOS POR SEMANA",             s_week,  r, palette) + 1
    r = write_table(ws, f"📅 {camara.upper()} — PROMEDIOS POR QUINCENA (15 días)", s_quin,  r, palette) + 1
    write_table(ws,     f"📅 {camara.upper()} — PROMEDIOS POR MES",                s_month, r, palette)


def build_general_sheet(ws):
    ws.title = "Resumen General"
    ws.sheet_view.showGridLines = False
    palette = PALETA["General"]

    periodos = [
        ("SEMANA",              "Semana"),
        ("QUINCENA (15 días)",  "Quincena"),
        ("MES",                 "Mes"),
    ]

    current_row = 2
    for periodo_label, periodo_key in periodos:
        # Calcular cuántas columnas ocupa cada cámara
        cam_widths = {c: len(summaries[c][periodo_key].columns) for c in CAMARAS}
        total_cols = sum(cam_widths.values()) + (len(CAMARAS) - 1)  # +gaps

        ws.merge_cells(
            start_row=current_row, start_column=1,
            end_row=current_row,   end_column=total_cols,
        )
        cell = ws.cell(row=current_row, column=1,
                       value=f"📅  PROMEDIOS POR {periodo_label}")
        cell.font      = make_font("FFFFFF", bold=True, size=12)
        cell.fill      = make_fill(palette["header"])
        cell.alignment = CENTER
        current_row += 1

        col_offset = 1
        max_rows   = 0
        for camara in CAMARAS:
            data   = summaries[camara][periodo_key]
            pal    = PALETA[camara]
            n_cols = cam_widths[camara]

            ws.merge_cells(
                start_row=current_row, start_column=col_offset,
                end_row=current_row,   end_column=col_offset + n_cols - 1,
            )
            cell = ws.cell(row=current_row, column=col_offset,
                           value=f"🐝 Cámara: {camara}")
            cell.font, cell.fill, cell.alignment = (
                make_font("FFFFFF", bold=True, size=10),
                make_fill(pal["subheader"]),
                CENTER,
            )

            for ci, col_name in enumerate(data.columns, col_offset):
                cell = ws.cell(row=current_row + 1, column=ci, value=col_name)
                cell.font, cell.fill, cell.alignment, cell.border = (
                    make_font("FFFFFF", bold=True, size=9),
                    make_fill(pal["subheader"]),
                    CENTER, BORDER,
                )

            white = make_fill("FFFFFF")
            alt   = make_fill(pal["alt"])
            for ri, (_, row) in enumerate(data.iterrows()):
                fill = alt if ri % 2 == 0 else white
                for ci, val in enumerate(row, col_offset):
                    cell = ws.cell(row=current_row + 2 + ri, column=ci, value=val)
                    cell.font, cell.fill, cell.alignment, cell.border = (
                        make_font(), fill, CENTER, BORDER,
                    )

            max_rows   = max(max_rows, len(data))
            col_offset += n_cols + 1   # gap entre cámaras

        current_row += 2 + max_rows + 3

    # Anchos de columna para el resumen general
    widths = []
    for camara in CAMARAS:
        n = len(summaries[camara]["Semana"].columns)
        widths += [22] + [19, 12] * ((n - 1) // 2 + 1)
        widths += [3]  # gap
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


# ─── Construir libro Excel ────────────────────────────────────────────────────
wb = Workbook()
build_general_sheet(wb.active)

for camara in CAMARAS:
    build_camera_sheet(
        wb, camara,
        summaries[camara]["Semana"],
        summaries[camara]["Quincena"],
        summaries[camara]["Mes"],
    )

wb.save(OUTPUT_XLS)
print(f"✅ Archivo guardado: {OUTPUT_XLS}")
print(f"   Hojas: {', '.join(wb.sheetnames)}")