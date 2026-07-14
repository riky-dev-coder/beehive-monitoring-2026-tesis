import pandas as pd
import numpy as np
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ─── Configuración ────────────────────────────────────────────────────────────
INPUT_CSV  = "feeds_1_.csv"
OUTPUT_XLS = "analisis_promedios.xlsx"

# Mapeo de columnas del CSV a variables de interés
COL_TEMP    = "field1"   # Temperatura  (°C)
COL_HUMEDAD = "field4"   # Humedad      (%)
COL_PESO    = "field7"   # Peso         (kg)

# ─── Carga y limpieza ─────────────────────────────────────────────────────────
df = pd.read_csv(INPUT_CSV, parse_dates=["created_at"])
df["created_at"] = pd.to_datetime(df["created_at"], utc=True).dt.tz_localize(None)

# Reemplazar 0 y nulos con NaN para excluirlos de los promedios
for col in [COL_TEMP, COL_HUMEDAD, COL_PESO]:
    df[col] = df[col].replace(0, np.nan)

df = df[["created_at", COL_TEMP, COL_HUMEDAD, COL_PESO]].copy()
df.columns = ["fecha", "temperatura", "humedad", "peso"]

# ─── Períodos ─────────────────────────────────────────────────────────────────
df["semana"]    = df["fecha"].dt.to_period("W")
df["quincena"]  = df["fecha"].apply(
    lambda x: f"{x.year}-{x.month:02d}-{'Q1 (1-15)' if x.day <= 15 else 'Q2 (16-fin)'}"
)
df["mes"] = df["fecha"].dt.to_period("M")

def make_summary(df, group_col, label):
    grp = df.groupby(group_col)[["temperatura", "humedad", "peso"]].agg(["mean", "count"])
    grp.columns = [
        "Temp. Promedio (°C)", "N_temp",
        "Humedad Promedio (%)", "N_humedad",
        "Peso Promedio (kg)",  "N_peso",
    ]
    grp = grp.reset_index()
    grp[label] = grp[group_col].astype(str)
    grp = grp.drop(columns=[group_col])
    grp[["Temp. Promedio (°C)", "Humedad Promedio (%)", "Peso Promedio (kg)"]] = (
        grp[["Temp. Promedio (°C)", "Humedad Promedio (%)", "Peso Promedio (kg)"]].round(2)
    )
    return grp[[label, "Temp. Promedio (°C)", "N_temp",
                "Humedad Promedio (%)", "N_humedad",
                "Peso Promedio (kg)", "N_peso"]]

weekly    = make_summary(df, "semana",   "Semana")
quincenal = make_summary(df, "quincena", "Quincena")
monthly   = make_summary(df, "mes",      "Mes")

# ─── Estilos Excel ────────────────────────────────────────────────────────────
HEADER_FILL = PatternFill("solid", start_color="1F4E79", end_color="1F4E79")
SUB_FILL    = PatternFill("solid", start_color="2E75B6", end_color="2E75B6")
ALT_FILL    = PatternFill("solid", start_color="D6E4F0", end_color="D6E4F0")
WHITE_FILL  = PatternFill("solid", start_color="FFFFFF", end_color="FFFFFF")
HEADER_FONT = Font(name="Arial", bold=True, color="FFFFFF", size=11)
SUB_FONT    = Font(name="Arial", bold=True, color="FFFFFF", size=10)
DATA_FONT   = Font(name="Arial", size=10)
CENTER      = Alignment(horizontal="center", vertical="center")
thin        = Side(style="thin", color="BFBFBF")
BORDER      = Border(left=thin, right=thin, top=thin, bottom=thin)

def write_table(ws, title, data, start_row):
    n_cols = len(data.columns)
    ws.merge_cells(
        start_row=start_row, start_column=1,
        end_row=start_row,   end_column=n_cols,
    )
    cell = ws.cell(row=start_row, column=1, value=title)
    cell.font, cell.fill, cell.alignment = HEADER_FONT, HEADER_FILL, CENTER
    start_row += 1

    for col_idx, col_name in enumerate(data.columns, 1):
        cell = ws.cell(row=start_row, column=col_idx, value=col_name)
        cell.font, cell.fill, cell.alignment, cell.border = SUB_FONT, SUB_FILL, CENTER, BORDER
    start_row += 1

    for row_idx, (_, row) in enumerate(data.iterrows()):
        fill = ALT_FILL if row_idx % 2 == 0 else WHITE_FILL
        for col_idx, val in enumerate(row, 1):
            cell = ws.cell(row=start_row + row_idx, column=col_idx, value=val)
            cell.font, cell.fill, cell.alignment, cell.border = DATA_FONT, fill, CENTER, BORDER

    return start_row + len(data) + 2

def setup_sheet(ws, title, w, q, m):
    ws.title = title
    ws.sheet_view.showGridLines = False
    for i, width in enumerate([22, 20, 14, 20, 14, 20, 14], 1):
        ws.column_dimensions[get_column_letter(i)].width = width
    r = 2
    r = write_table(ws, "📅 PROMEDIOS POR SEMANA",                    w, r) + 1
    r = write_table(ws, "📅 PROMEDIOS POR QUINCENA (cada 15 días)",   q, r) + 1
    write_table(ws,     "📅 PROMEDIOS POR MES",                       m, r)

def single_metric_sheet(wb, name, col, n_col, w, q, m):
    ws = wb.create_sheet(name)
    ws.sheet_view.showGridLines = False
    for i, width in enumerate([24, 22, 16], 1):
        ws.column_dimensions[get_column_letter(i)].width = width

    def extract(df, period): return df[[period, col, n_col]].copy()

    r = 2
    r = write_table(ws, f"📅 {name.upper()} - POR SEMANA",    extract(w, "Semana"),   r) + 1
    r = write_table(ws, f"📅 {name.upper()} - POR QUINCENA",  extract(q, "Quincena"), r) + 1
    write_table(ws,     f"📅 {name.upper()} - POR MES",       extract(m, "Mes"),      r)

# ─── Construcción del libro ───────────────────────────────────────────────────
wb = Workbook()
setup_sheet(wb.active, "Resumen General", weekly, quincenal, monthly)
single_metric_sheet(wb, "Temperatura", "Temp. Promedio (°C)",  "N_temp",    weekly, quincenal, monthly)
single_metric_sheet(wb, "Humedad",     "Humedad Promedio (%)", "N_humedad", weekly, quincenal, monthly)
single_metric_sheet(wb, "Peso",        "Peso Promedio (kg)",   "N_peso",    weekly, quincenal, monthly)

wb.save(OUTPUT_XLS)
print(f"✅ Archivo guardado: {OUTPUT_XLS}")